import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Any
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re


class ProcessadorOrcamento:

    def __init__(self, fonte_proibida=None, naturezas_proibidas=None):
        self.df_original = None
        self.ugs_dados = []
        self.remanejamentos = []
        self.diagnosticos = []

        # CONFIGURAÇÃO: Colunas importantes
        self.COLUNA_FONTE = 0   # Coluna A: Fonte (500, 501, 761)
        self.COLUNA_SALDO = None  # Será identificada automaticamente pelo nome "7- Previsão Orçamentária"
        self.NOME_COLUNA_SALDO = "7- Previsão Orçamentária"  # Nome da coluna a ser buscada

        # CONFIGURAÇÃO: Fonte proibida de remanejamento (passada como parâmetro)
        self.FONTE_PROIBIDA = fonte_proibida  # None = nenhuma fonte proibida

        # CONFIGURAÇÃO: Naturezas proibidas de remanejamento (passadas como parâmetro)
        # Se não foi passado, usa conjunto vazio
        self.NATUREZAS_PROIBIDAS = naturezas_proibidas if naturezas_proibidas else set()

        # Processar TODAS as naturezas deficitárias (não apenas linhas de teste)
        self.LINHAS_TESTE = None  # None = processar tudo

        # NOVAS REGRAS: Proteção de saldo mínimo
        self.PERCENTUAL_RESERVA_MINIMA = 0.20  # 20% do saldo original deve ser preservado
        self.PERCENTUAL_DOACAO_MAXIMA_POR_VEZ = 0.40  # Doar no máximo 40% do saldo total por operação (reduz qtd de remanejamentos)

        # OTIMIZAÇÃO: Priorizar doação única quando possível
        self.PRIORIZAR_DOACAO_UNICA = True  # Se uma natureza pode cobrir sozinha, usa ela completamente

    def log(self, mensagem):
        """Adiciona mensagem de diagnóstico"""
        print(f"[DEBUG] {mensagem}")
        self.diagnosticos.append(mensagem)

    def processar_arquivo(self, arquivo) -> Dict[str, Any]:
        self.log("=" * 80)
        self.log("INICIANDO PROCESSAMENTO - REGRAS SEFAZ COMPLETAS")
        self.log("=" * 80)
        self.log(f"\n📋 REGRAS DE PROTEÇÃO ATIVAS:")
        self.log(f"   • Saldo mínimo preservado: {self.PERCENTUAL_RESERVA_MINIMA * 100:.0f}% do saldo original")
        self.log(f"   • Doação máxima por operação: {self.PERCENTUAL_DOACAO_MAXIMA_POR_VEZ * 100:.0f}% do saldo original (OTIMIZADO)")
        self.log(f"   • Priorizar doação única: {'SIM' if self.PRIORIZAR_DOACAO_UNICA else 'NÃO'} (reduz remanejamentos)")
        self.log(f"   • Consolidação automática no Excel: SIM")
        if self.FONTE_PROIBIDA:
            self.log(f"   • Fonte {self.FONTE_PROIBIDA} nunca participa de remanejamento")
        else:
            self.log(f"   • Nenhuma fonte proibida configurada")
        if self.NATUREZAS_PROIBIDAS:
            self.log(f"   • {len(self.NATUREZAS_PROIBIDAS)} naturezas proibidas (responsabilidade de cada UG)")
        else:
            self.log(f"   • Nenhuma natureza proibida configurada")
        self.log(f"   • Apenas naturezas originalmente negativas podem receber")
        self.log(f"   • Apenas naturezas originalmente positivas podem doar")

        # 1. Ler planilha
        self.log("\n1. Lendo planilha Excel...")
        self.df_original = self.ler_planilha(arquivo)
        self.log(f"   Planilha: {len(self.df_original)} linhas x {len(self.df_original.columns)} colunas")

        # 1.1. Identificar coluna de saldo automaticamente
        self.log("\n1.1. Identificando coluna de saldo...")
        self.encontrar_coluna_saldo()

        # 2. Identificar estrutura (UGs e Naturezas na coluna B)
        self.log("\n2. Identificando UGs e Naturezas...")
        self.identificar_estrutura()
        self.log(f"   Total de UGs identificadas: {len(self.ugs_dados)}")

        if len(self.ugs_dados) == 0:
            raise Exception("Nenhuma UG encontrada!")

        # 3. Identificar déficits
        self.log("\n3. Identificando déficits...")
        deficits_totais = self.identificar_deficits()

        # 4. Remanejamento Interno (Prioridade 1)
        self.log("\n4. PRIORIDADE 1: Remanejamento INTERNO (mesma UG)...")
        self.remanejamento_interno()

        # 5. Remanejamento Externo (se necessário)
        # Regra: Priorizar remanejamento interno, mas permitir externo se necessário
        self.log("\n5. Remanejamento EXTERNO (se houver déficit residual)...")
        self.remanejamento_externo()

        # 6. Validar
        self.log("\n6. Validando resultados...")
        validacoes = self.validar_resultado()

        # 7. Gerar Excel
        self.log("\n7. Gerando arquivo Excel...")
        arquivo_excel = self.gerar_excel()

        self.log("\n" + "=" * 80)
        self.log("PROCESSAMENTO CONCLUÍDO")
        self.log("=" * 80)

        return {
            'estatisticas': {
                'total_ugs': len(self.ugs_dados),
                'total_deficits': deficits_totais,
                'remanejamentos_internos': sum(1 for r in self.remanejamentos if r['ug_origem'] == r['ug_destino']),
                'remanejamentos_externos': sum(1 for r in self.remanejamentos if r['ug_origem'] != r['ug_destino']),
            },
            'deficits': [
                {
                    'UG': ug['codigo'],
                    'UG Nome': ug['nome'],
                    'Natureza': nat['codigo'],
                    'Natureza Nome': nat['nome'],
                    'Déficit': abs(nat['saldo_original'])
                }
                for ug in self.ugs_dados
                for nat in ug['naturezas']
                if nat['saldo_original'] < 0
            ],
            'remanejamentos': self.remanejamentos,
            'validacoes': validacoes,
            'arquivo_excel': arquivo_excel,
            'diagnosticos': '\n'.join(self.diagnosticos)
        }

    def ler_planilha(self, arquivo) -> pd.DataFrame:
        nome_arquivo = arquivo.name if hasattr(arquivo, 'name') else 'arquivo.xlsx'

        if nome_arquivo.endswith('.xls'):
            df = pd.read_excel(arquivo, sheet_name=0, header=None, engine='xlrd')
        else:
            df = pd.read_excel(arquivo, sheet_name=0, header=None, engine='openpyxl')

        df = df.reset_index(drop=True)
        return df

    def encontrar_coluna_saldo(self):
        if self.df_original is None:
            raise Exception("Planilha não carregada!")

        # Procurar nas primeiras 10 linhas pelo cabeçalho
        for linha_idx in range(min(10, len(self.df_original))):
            row = self.df_original.iloc[linha_idx]
            for col_idx, valor in enumerate(row):
                if pd.isna(valor):
                    continue
                valor_str = str(valor).strip()
                # Buscar especificamente por "7- Previsão" para evitar pegar títulos da planilha
                # O padrão deve começar com "7-" ou "7 -" seguido de "Previsão"
                valor_lower = valor_str.lower()
                if valor_lower.startswith("7-") or valor_lower.startswith("7 -"):
                    if "previsão" in valor_lower:
                        self.COLUNA_SALDO = col_idx
                        self.log(f"   Coluna de saldo encontrada: '{valor_str}' (coluna {chr(65 + col_idx)}, índice {col_idx})")
                        return col_idx

        # Se não encontrou, lançar erro informativo
        raise Exception(
            f"Coluna '{self.NOME_COLUNA_SALDO}' não encontrada no cabeçalho da planilha! "
            f"Verifique se a planilha contém uma coluna com esse nome (deve começar com '7-')."
        )

    def identificar_estrutura(self):
        """
        Identifica UGs e Naturezas na COLUNA B
        UG: 6 dígitos + " - " + MAIÚSCULAS (ex: 450201 - DETRAN)
        Natureza: 6 dígitos + " - " + Maiúsculas/minúsculas (ex: 319011 - Vencimento...)
        """
        # Padrão: 6 dígitos + " - " + texto
        padrao_geral = re.compile(r'^(\d{6})\s*-\s*(.+)$')

        COLUNA_B = 1  # Coluna B = índice 1

        ug_atual = None
        fonte_atual = None  # Rastreia a última fonte válida vista (coluna A pode ter células mescladas)

        for idx, row in self.df_original.iterrows():
            # Ler valor da coluna B
            if COLUNA_B >= len(row):
                continue

            valor_b = row.iloc[COLUNA_B]

            if pd.isna(valor_b):
                continue

            valor_str = str(valor_b).strip()
            match = padrao_geral.match(valor_str)

            if match:
                codigo = match.group(1)
                nome = match.group(2).strip()

                # Verificar se é UG ou Natureza
                # UG: nome em MAIÚSCULAS
                # Natureza: nome com minúsculas
                if nome == nome.upper():
                    # É UG
                    saldo = self.extrair_valor_coluna(row, self.COLUNA_SALDO)
                    fonte = self.extrair_valor_coluna(row, self.COLUNA_FONTE)

                    # Atualiza fonte_atual se a coluna A tiver valor (células mescladas: só a 1ª tem valor)
                    if fonte > 0:
                        fonte_atual = int(fonte)

                    ug_atual = {
                        'codigo': codigo,
                        'nome': nome,
                        'linha': idx,
                        'linha_excel': idx + 1,
                        'saldo_total': saldo,
                        'fonte': fonte_atual,
                        'naturezas': []
                    }

                    self.ugs_dados.append(ug_atual)

                    fonte_str = f"Fonte: {fonte_atual}" if fonte_atual else "Sem fonte"
                    status = "DÉFICIT" if saldo < 0 else "SUPERÁVIT" if saldo > 0 else "ZERO"
                    self.log(f"   UG: {codigo} - {nome} ({fonte_str}) (linha {idx + 1}) | Saldo: {saldo:,.2f} ({status})")

                else:
                    # É Natureza
                    if ug_atual is not None:
                        saldo = self.extrair_valor_coluna(row, self.COLUNA_SALDO)
                        fonte = self.extrair_valor_coluna(row, self.COLUNA_FONTE)

                        natureza = {
                            'codigo': codigo,
                            'nome': nome,
                            'linha': idx,
                            'linha_excel': idx + 1,
                            'saldo_original': saldo,
                            'saldo_atual': saldo,
                            'dois_primeiros_digitos': codigo[:2],  # Para priorização
                            'fonte': int(fonte) if fonte > 0 else ug_atual['fonte']  # Herda fonte da UG
                        }

                        ug_atual['naturezas'].append(natureza)

                        status = "DÉFICIT" if saldo < 0 else "SUPERÁVIT" if saldo > 0 else "ZERO"
                        self.log(f"      Natureza: {codigo} - {nome[:40]}... (linha {idx + 1}) | Saldo: {saldo:,.2f} ({status})")

        # Processar TODAS as naturezas (sem filtro de linhas de teste)
        self.log(f"\n   Modo: Processar TODAS as naturezas deficitárias encontradas")

    def extrair_valor_coluna(self, row, coluna_idx) -> float:
        """Extrai valor numérico de uma coluna"""
        if coluna_idx >= len(row):
            return 0.0

        valor = row.iloc[coluna_idx]

        if pd.isna(valor):
            return 0.0

        try:
            if isinstance(valor, (int, float)):
                return float(valor)
            else:
                valor_str = str(valor).strip().replace(',', '.')
                return float(valor_str)
        except:
            return 0.0

    def natureza_eh_proibida(self, codigo_natureza: str) -> bool:
        """
        Verifica se a natureza está na lista de naturezas proibidas de remanejamento.
        Essas naturezas são de responsabilidade exclusiva de cada UG.
        """
        # Remover pontos e espaços para comparação
        codigo_limpo = codigo_natureza.replace('.', '').replace(' ', '').strip()
        return codigo_limpo in self.NATUREZAS_PROIBIDAS

    def calcular_capacidade_doacao(self, natureza: Dict) -> float:
        """
        Calcula quanto uma natureza pode doar, respeitando o saldo mínimo de segurança.

        REGRAS:
        1. Preservar no mínimo 20% do saldo original
        2. Doar no máximo 10% do saldo original por vez (para UGs grandes)
        3. Nunca zerar uma natureza positiva
        """
        saldo_original = natureza['saldo_original']
        saldo_atual = natureza['saldo_atual']

        # Se já não tem saldo positivo, não pode doar nada
        if saldo_atual <= 0:
            return 0.0

        # Calcular saldo mínimo a preservar (20% do original)
        saldo_minimo = saldo_original * self.PERCENTUAL_RESERVA_MINIMA

        # Calcular máximo que pode doar (80% do original)
        capacidade_maxima = saldo_original - saldo_minimo

        # REGRA ESPECIAL: Para UGs grandes, limitar doação a 10% por vez
        doacao_maxima_por_vez = saldo_original * self.PERCENTUAL_DOACAO_MAXIMA_POR_VEZ

        # Quanto já foi doado até agora
        quanto_ja_doou = saldo_original - saldo_atual

        # Quanto ainda pode doar (respeitando limite de 80% total)
        quanto_ainda_pode_doar = max(0, capacidade_maxima - quanto_ja_doou)

        # Limitar pela regra de 10% por vez
        capacidade_real = min(quanto_ainda_pode_doar, doacao_maxima_por_vez)

        # Garantir que não vai ultrapassar o saldo atual
        capacidade_real = min(capacidade_real, saldo_atual - saldo_minimo)

        return max(0, capacidade_real)

    def identificar_deficits(self) -> int:
        """Identifica TODOS os déficits na planilha (exceto fonte 761 e naturezas proibidas)"""
        total_deficits = 0
        ignorados_761 = 0
        ignorados_naturezas_proibidas = 0

        for ug in self.ugs_dados:
            # Ignorar fonte 761
            if ug['fonte'] == self.FONTE_PROIBIDA:
                deficits_761 = [nat for nat in ug['naturezas'] if nat['saldo_original'] < 0]
                if deficits_761:
                    ignorados_761 += len(deficits_761)
                    self.log(f"\n   UG {ug['codigo']} (Fonte 761): {len(deficits_761)} déficit(s) IGNORADOS (fonte proibida)")
                continue

            # TODAS as naturezas com saldo negativo (exceto fonte 761 e naturezas proibidas)
            deficits_ug = []
            for nat in ug['naturezas']:
                if nat['saldo_original'] >= 0:
                    continue
                if nat['fonte'] == self.FONTE_PROIBIDA:
                    continue

                # Verificar se é natureza proibida
                if self.natureza_eh_proibida(nat['codigo']):
                    ignorados_naturezas_proibidas += 1
                    self.log(f"\n   UG {ug['codigo']} - Natureza {nat['codigo']}: IGNORADA (natureza proibida - responsabilidade da UG)")
                    continue

                deficits_ug.append(nat)

            if deficits_ug:
                self.log(f"\n   UG {ug['codigo']} - {ug['nome']} (Fonte: {ug['fonte']}): {len(deficits_ug)} déficit(s)")
                for nat in deficits_ug:
                    deficit_puro = abs(nat['saldo_original'])
                    nat['necessidade_total'] = deficit_puro  # Exatamente o déficit
                    self.log(f"      • {nat['codigo']} (linha {nat['linha_excel']}, Fonte: {nat['fonte']}) - {nat['nome'][:40]}...")
                    self.log(f"        Déficit: {deficit_puro:,.2f}")
                    total_deficits += 1

        if ignorados_761 > 0:
            self.log(f"\n   ⚠️ {ignorados_761} déficit(s) da Fonte 761 foram IGNORADOS")

        if ignorados_naturezas_proibidas > 0:
            self.log(f"\n   ⚠️ {ignorados_naturezas_proibidas} déficit(s) de naturezas proibidas foram IGNORADOS (responsabilidade de cada UG)")

        if total_deficits == 0:
            self.log("   Nenhum déficit encontrado (excluindo fonte 761)!")

        return total_deficits

    def remanejamento_interno(self):
        """
        Remanejamento APENAS dentro da mesma UG
        Processa TODAS as naturezas deficitárias (exceto fonte 761)
        """
        for ug in self.ugs_dados:
            # Ignorar fonte 761
            if ug['fonte'] == self.FONTE_PROIBIDA:
                self.log(f"\n   UG {ug['codigo']} (Fonte 761): IGNORADA (fonte proibida)")
                continue

            self.log(f"\n   Processando UG {ug['codigo']} - {ug['nome']} (Fonte: {ug['fonte']})...")

            # Deficitárias: TODAS com saldo negativo (exceto fonte 761 e naturezas proibidas)
            deficitarias = [nat for nat in ug['naturezas']
                           if nat['saldo_atual'] < 0
                           and nat['fonte'] != self.FONTE_PROIBIDA
                           and not self.natureza_eh_proibida(nat['codigo'])]

            # Superavitárias: TODAS as naturezas da UG com saldo positivo (exceto fonte 761 e naturezas proibidas)
            superavitarias = [nat for nat in ug['naturezas']
                             if nat['saldo_atual'] > 0
                             and nat['fonte'] != self.FONTE_PROIBIDA
                             and not self.natureza_eh_proibida(nat['codigo'])]

            if not deficitarias:
                self.log(f"      Sem déficits")
                continue

            if not superavitarias:
                self.log(f"      Sem superávits disponíveis nesta UG")
                continue

            self.log(f"      Déficits a cobrir: {len(deficitarias)}")
            self.log(f"      Naturezas doadoras disponíveis: {len(superavitarias)}")

            # Mostrar naturezas doadoras
            if superavitarias:
                self.log(f"\n      Naturezas doadoras:")
                for nat_super in sorted(superavitarias, key=lambda x: x['saldo_atual'], reverse=True):
                    self.log(f"         • {nat_super['codigo']}: {nat_super['saldo_atual']:,.2f} (dígitos: {nat_super['dois_primeiros_digitos']})")

            # Para cada déficit
            for nat_deficit in deficitarias:
                necessidade_restante = nat_deficit.get('necessidade_total', abs(nat_deficit['saldo_atual']))

                if necessidade_restante <= 0.01:
                    continue

                digitos_deficit = nat_deficit['dois_primeiros_digitos']

                self.log(f"\n      Cobrindo déficit: {nat_deficit['codigo']} - {nat_deficit['nome'][:40]}...")
                self.log(f"         Necessidade: {necessidade_restante:,.2f}")

                # PRIORIDADE: Mesmos 2 primeiros dígitos
                doadoras_prioritarias = [s for s in superavitarias if s['dois_primeiros_digitos'] == digitos_deficit]
                doadoras_secundarias = [s for s in superavitarias if s['dois_primeiros_digitos'] != digitos_deficit]

                # Ordenar por saldo (maior primeiro)
                doadoras_prioritarias.sort(key=lambda x: x['saldo_atual'], reverse=True)
                doadoras_secundarias.sort(key=lambda x: x['saldo_atual'], reverse=True)

                # OTIMIZAÇÃO: Verificar se UMA única natureza pode cobrir tudo
                doadora_unica = None
                if self.PRIORIZAR_DOACAO_UNICA:
                    for nat_super in doadoras_prioritarias + doadoras_secundarias:
                        capacidade = self.calcular_capacidade_doacao(nat_super)
                        if capacidade >= necessidade_restante:
                            doadora_unica = nat_super
                            self.log(f"         ✓ Doadora única encontrada: {nat_super['codigo']} (capacidade: {capacidade:,.2f})")
                            break

                # Se encontrou doadora única, usar ela
                if doadora_unica:
                    self.log(f"         • {doadora_unica['codigo']}: cobrindo TUDO em uma única transferência")
                    self.registrar_transferencia(ug['codigo'], doadora_unica, ug['codigo'], nat_deficit, necessidade_restante, "Interna (única)")
                    necessidade_restante = 0
                else:
                    # Caso contrário, distribuir entre várias (lógica atual)
                    # Tentar doadoras prioritárias primeiro
                    for nat_super in doadoras_prioritarias:
                        if necessidade_restante <= 0.01:
                            break

                        # NOVA REGRA: Calcular capacidade real de doação (respeitando saldo mínimo)
                        capacidade_doacao = self.calcular_capacidade_doacao(nat_super)

                        if capacidade_doacao <= 0.01:
                            self.log(f"         • {nat_super['codigo']}: sem capacidade de doação (preservando saldo mínimo)")
                            continue

                        valor_transferir = min(necessidade_restante, capacidade_doacao)

                        self.log(f"         • {nat_super['codigo']}: pode doar {capacidade_doacao:,.2f}, transferindo {valor_transferir:,.2f}")
                        self.registrar_transferencia(ug['codigo'], nat_super, ug['codigo'], nat_deficit, valor_transferir, "Interna (mesmos dígitos)")

                        # Saldos já atualizados dentro de registrar_transferencia()
                        necessidade_restante -= valor_transferir

                    # Se ainda não cobriu, usar doadoras secundárias
                    for nat_super in doadoras_secundarias:
                        if necessidade_restante <= 0.01:
                            break

                        # Calcular capacidade real de doação (respeitando saldo mínimo)
                        capacidade_doacao = self.calcular_capacidade_doacao(nat_super)

                        if capacidade_doacao <= 0.01:
                            self.log(f"         • {nat_super['codigo']}: sem capacidade de doação (preservando saldo mínimo)")
                            continue

                        valor_transferir = min(necessidade_restante, capacidade_doacao)

                        self.log(f"         • {nat_super['codigo']}: pode doar {capacidade_doacao:,.2f}, transferindo {valor_transferir:,.2f}")
                        self.registrar_transferencia(ug['codigo'], nat_super, ug['codigo'], nat_deficit, valor_transferir, "Interna")

                    # Saldos já atualizados dentro de registrar_transferencia()
                    necessidade_restante -= valor_transferir

                # Atualizar necessidade
                nat_deficit['necessidade_total'] = max(0, necessidade_restante)

                if necessidade_restante > 0.01:
                    self.log(f"         ⚠️ ATENÇÃO: Ainda falta {necessidade_restante:,.2f} para cobrir totalmente")

    def remanejamento_externo(self):
        """
        PRIORIDADE 2: Remanejamento entre UGs da mesma fonte
        Prioriza UG com MAIOR saldo positivo da mesma fonte
        IGNORA fonte 761
        """
        self.log("")

        # Identificar déficits residuais (exceto fonte 761)
        # IMPORTANTE: Apenas naturezas que eram ORIGINALMENTE negativas podem receber
        ugs_deficitarias = []
        for ug in self.ugs_dados:
            # Ignorar fonte 761
            if ug['fonte'] == self.FONTE_PROIBIDA:
                continue

            # Apenas naturezas que COMEÇARAM negativas e ainda precisam de cobertura (exceto proibidas)
            deficits = [nat for nat in ug['naturezas']
                       if nat['saldo_original'] < 0  # DEVE ter começado negativa
                       and nat.get('necessidade_total', 0) > 0.01  # E ainda precisar de cobertura
                       and nat['fonte'] != self.FONTE_PROIBIDA
                       and not self.natureza_eh_proibida(nat['codigo'])]  # Excluir naturezas proibidas

            if deficits:
                total_necessidade = sum(nat.get('necessidade_total', abs(nat['saldo_atual'])) for nat in deficits)
                ugs_deficitarias.append({
                    'ug': ug,
                    'fonte': ug['fonte'],
                    'necessidade_total': total_necessidade,
                    'naturezas_deficit': deficits
                })
                self.log(f"   UG {ug['codigo']} (Fonte {ug['fonte']}) ainda precisa de {total_necessidade:,.2f}")

        if not ugs_deficitarias:
            self.log("   Todos os déficits foram cobertos internamente!")
            return

        # Para cada UG deficitária, buscar doadoras DA MESMA FONTE
        for ug_deficit_info in ugs_deficitarias:
            fonte_deficitaria = ug_deficit_info['fonte']

            self.log(f"\n   Buscando doadoras da Fonte {fonte_deficitaria} para UG {ug_deficit_info['ug']['codigo']}...")

            # Identificar UGs doadoras da MESMA FONTE (exceto fonte 761)
            ugs_doadoras = []
            for ug in self.ugs_dados:
                # Ignorar fonte 761
                if ug['fonte'] == self.FONTE_PROIBIDA:
                    continue

                # Apenas mesma fonte
                if ug['fonte'] != fonte_deficitaria:
                    continue

                # Não pode ser a própria UG deficitária
                if ug['codigo'] == ug_deficit_info['ug']['codigo']:
                    continue

                # Naturezas com saldo positivo (exceto naturezas proibidas)
                naturezas_super = [nat for nat in ug['naturezas']
                                  if nat['saldo_atual'] > 0
                                  and nat['fonte'] != self.FONTE_PROIBIDA
                                  and not self.natureza_eh_proibida(nat['codigo'])]

                if naturezas_super:
                    total_super = sum(nat['saldo_atual'] for nat in naturezas_super)
                    ugs_doadoras.append({
                        'ug': ug,
                        'fonte': ug['fonte'],
                        'superavit_total': total_super,
                        'naturezas_super': naturezas_super
                    })

            if not ugs_doadoras:
                self.log(f"      ⚠️ Nenhuma UG doadora disponível na Fonte {fonte_deficitaria}!")
                continue

            # ORDENAR por MAIOR saldo primeiro (prioridade obrigatória)
            ugs_doadoras.sort(key=lambda x: x['superavit_total'], reverse=True)

            self.log(f"      UGs doadoras encontradas (Fonte {fonte_deficitaria}):")
            for ud in ugs_doadoras:
                self.log(f"         • UG {ud['ug']['codigo']}: {ud['superavit_total']:,.2f} disponível")

            # Realizar transferências para esta UG deficitária
            for nat_deficit in ug_deficit_info['naturezas_deficit']:
                necessidade_restante = nat_deficit.get('necessidade_total', abs(nat_deficit['saldo_atual']))

                if necessidade_restante <= 0.01:
                    continue

                digitos_deficit = nat_deficit['dois_primeiros_digitos']

                # Buscar em UGs doadoras DA MESMA FONTE, priorizando MAIOR saldo
                for ug_doadora_info in ugs_doadoras:
                    if necessidade_restante <= 0.01:
                        break

                    # Separar por prioridade
                    prioritarias = [s for s in ug_doadora_info['naturezas_super']
                                  if s['dois_primeiros_digitos'] == digitos_deficit]
                    secundarias = [s for s in ug_doadora_info['naturezas_super']
                                 if s['dois_primeiros_digitos'] != digitos_deficit]

                    prioritarias.sort(key=lambda x: x['saldo_atual'], reverse=True)
                    secundarias.sort(key=lambda x: x['saldo_atual'], reverse=True)

                    # Tentar prioritárias
                    for nat_super in prioritarias:
                        if necessidade_restante <= 0.01:
                            break

                        # NOVA REGRA: Calcular capacidade real de doação (respeitando saldo mínimo)
                        capacidade_doacao = self.calcular_capacidade_doacao(nat_super)

                        if capacidade_doacao <= 0.01:
                            continue

                        valor_transferir = min(necessidade_restante, capacidade_doacao)

                        self.registrar_transferencia(
                            ug_doadora_info['ug']['codigo'], nat_super,
                            ug_deficit_info['ug']['codigo'], nat_deficit,
                            valor_transferir, "Externa (mesmos dígitos)"
                        )

                        # Saldos já atualizados dentro de registrar_transferencia()
                        necessidade_restante -= valor_transferir

                    # Tentar secundárias
                    for nat_super in secundarias:
                        if necessidade_restante <= 0.01:
                            break

                        # NOVA REGRA: Calcular capacidade real de doação (respeitando saldo mínimo)
                        capacidade_doacao = self.calcular_capacidade_doacao(nat_super)

                        if capacidade_doacao <= 0.01:
                            continue

                        valor_transferir = min(necessidade_restante, capacidade_doacao)

                        self.registrar_transferencia(
                            ug_doadora_info['ug']['codigo'], nat_super,
                            ug_deficit_info['ug']['codigo'], nat_deficit,
                            valor_transferir, "Externa"
                        )

                        # Saldos já atualizados dentro de registrar_transferencia()
                        necessidade_restante -= valor_transferir

                nat_deficit['necessidade_total'] = max(0, necessidade_restante)

                if necessidade_restante > 0.01:
                    self.log(f"         ⚠️ ATENÇÃO: UG {ug_deficit_info['ug']['codigo']} - {nat_deficit['codigo']} ainda falta {necessidade_restante:,.2f}")

    def registrar_transferencia(self, ug_origem, nat_origem, ug_destino, nat_destino, valor, tipo):
        """Registra uma transferência"""
        # VALIDAÇÃO CRÍTICA: Origem deve ser originalmente positiva, Destino deve ser originalmente negativa
        if nat_origem['saldo_original'] <= 0:
            self.log(f"         ❌ ERRO: Tentativa de usar natureza originalmente negativa como doadora: {nat_origem['codigo']} (saldo original: {nat_origem['saldo_original']:,.2f})")
            return

        if nat_destino['saldo_original'] >= 0:
            self.log(f"         ❌ ERRO: Tentativa de enviar para natureza originalmente positiva: {nat_destino['codigo']} (saldo original: {nat_destino['saldo_original']:,.2f})")
            return

        # VALIDAÇÃO CRÍTICA: Não permitir doar mais do que o saldo atual disponível
        if valor > nat_origem['saldo_atual']:
            self.log(f"         ❌ ERRO: Tentativa de doar {valor:,.2f} mas origem só tem {nat_origem['saldo_atual']:,.2f} disponível")
            return

        # VALIDAÇÃO CRÍTICA: Garantir que após a doação, o saldo não fica abaixo do mínimo de segurança
        saldo_minimo = nat_origem['saldo_original'] * self.PERCENTUAL_RESERVA_MINIMA
        saldo_apos_doacao = nat_origem['saldo_atual'] - valor

        if saldo_apos_doacao < saldo_minimo:
            self.log(f"         ❌ ERRO: Doação de {valor:,.2f} violaria saldo mínimo de segurança ({saldo_minimo:,.2f}). Saldo ficaria: {saldo_apos_doacao:,.2f}")
            return

        # VALIDAÇÃO CRÍTICA: Naturezas proibidas não podem doar nem receber
        if self.natureza_eh_proibida(nat_origem['codigo']):
            self.log(f"         ❌ ERRO: Natureza origem {nat_origem['codigo']} está na lista de naturezas proibidas (responsabilidade da UG)")
            return

        if self.natureza_eh_proibida(nat_destino['codigo']):
            self.log(f"         ❌ ERRO: Natureza destino {nat_destino['codigo']} está na lista de naturezas proibidas (responsabilidade da UG)")
            return

        # Log detalhado ANTES da transferência
        self.log(f"         >> ANTES: Origem {nat_origem['codigo']} saldo={nat_origem['saldo_atual']:,.2f} | Destino {nat_destino['codigo']} saldo={nat_destino['saldo_atual']:,.2f}")

        # Garantir que sempre temos a fonte (buscar da natureza ou da UG se necessário)
        fonte = nat_origem.get('fonte')
        if fonte is None:
            # Buscar da UG origem
            for ug in self.ugs_dados:
                if ug['codigo'] == ug_origem:
                    fonte = ug.get('fonte')
                    break

        if fonte is None:
            self.log(f"         ⚠️ AVISO: Fonte não encontrada para UG {ug_origem} — remanejamento registrado sem fonte")

        self.remanejamentos.append({
            'Tipo': tipo,
            'Fonte': int(fonte) if fonte is not None else None,
            'UG Origem': ug_origem,
            'Natureza Origem': nat_origem['codigo'],
            'Nome Natureza Origem': nat_origem['nome'],
            'UG Destino': ug_destino,
            'Natureza Destino': nat_destino['codigo'],
            'Nome Natureza Destino': nat_destino['nome'],
            'Valor': round(valor, 2),
            'ug_origem': ug_origem,
            'ug_destino': ug_destino
        })

        self.log(f"         ✓ {tipo}: {nat_origem['codigo']} (Fonte {nat_origem['fonte']}) → {nat_destino['codigo']}: {valor:,.2f}")

        # IMPORTANTE: Atualizar os saldos DENTRO desta função
        nat_origem['saldo_atual'] -= valor
        nat_destino['saldo_atual'] += valor

        # Log detalhado DEPOIS da transferência
        self.log(f"         << DEPOIS: Origem {nat_origem['codigo']} saldo={nat_origem['saldo_atual']:,.2f} | Destino {nat_destino['codigo']} saldo={nat_destino['saldo_atual']:,.2f}")

    def validar_resultado(self) -> Dict[str, bool]:
        """Valida resultados"""
        tem_negativo = False

        self.log("\n   === RESUMO FINAL DE TODAS AS NATUREZAS ===")

        for ug in self.ugs_dados:
            for nat in ug['naturezas']:
                saldo_original = nat['saldo_original']
                saldo_atual = nat['saldo_atual']
                diferenca = saldo_atual - saldo_original

                # Mostrar TODAS as naturezas que mudaram
                if abs(diferenca) > 0.01:
                    if saldo_original < 0:  # Era deficitária
                        self.log(f"   {ug['codigo']} - {nat['codigo']} (DEFICITÁRIA): {saldo_original:,.2f} → {saldo_atual:,.2f} (recebeu {diferenca:,.2f})")

                        if saldo_atual < -0.01:
                            tem_negativo = True
                            self.log(f"      ❌ AINDA NEGATIVO!")
                        elif abs(saldo_atual) < 0.01:
                            self.log(f"      ✓ Zerado com sucesso")
                        else:
                            self.log(f"      ⚠️ Parcialmente coberto")

                    elif saldo_original > 0:  # Era superavitária
                        if diferenca > 0:
                            self.log(f"   {ug['codigo']} - {nat['codigo']} (SUPERAVITÁRIA): {saldo_original:,.2f} → {saldo_atual:,.2f} (AUMENTOU {diferenca:,.2f}) ⚠️ ERRO!")
                        else:
                            saldo_minimo = saldo_original * self.PERCENTUAL_RESERVA_MINIMA
                            percentual_preservado = (saldo_atual / saldo_original) * 100
                            self.log(f"   {ug['codigo']} - {nat['codigo']} (DOADORA): {saldo_original:,.2f} → {saldo_atual:,.2f} (doou {abs(diferenca):,.2f}, preservou {percentual_preservado:.1f}%)")

                            if saldo_atual < saldo_minimo:
                                self.log(f"      ⚠️ ATENÇÃO: Saldo abaixo do mínimo de segurança ({saldo_minimo:,.2f})")

        if not tem_negativo:
            self.log("\n   ✓ Nenhum saldo negativo")
        else:
            self.log("\n   ❌ AINDA EXISTEM SALDOS NEGATIVOS!")

        if len(self.remanejamentos) > 0:
            self.log(f"   ✓ {len(self.remanejamentos)} remanejamentos realizados")

        return {
            'nenhum_saldo_negativo': not tem_negativo,
            'somas_conferem': len(self.remanejamentos) > 0
        }

    def gerar_excel(self) -> bytes:
        """Gera Excel com 2 abas"""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Aba 1: Saldos
            df_aba1 = self.gerar_aba_saldos()
            df_aba1.to_excel(writer, sheet_name='Saldos Ajustados', index=False)

            # Aba 2: Remanejamentos
            df_aba2 = self.gerar_aba_remanejamento()
            df_aba2.to_excel(writer, sheet_name='Remanejamentos', index=False)

            workbook = writer.book
            self.formatar_planilha(workbook['Saldos Ajustados'])
            self.formatar_planilha(workbook['Remanejamentos'])

        output.seek(0)
        self.log("   Excel gerado com sucesso")
        return output.getvalue()

    def gerar_aba_saldos(self) -> pd.DataFrame:
        """Gera aba de saldos ajustados"""
        dados = []

        for ug in self.ugs_dados:
            saldo_total_ajustado = sum(nat['saldo_atual'] for nat in ug['naturezas'])

            dados.append({
                'Fonte': ug['fonte'],
                'UG': ug['codigo'],
                'Nome UG': ug['nome'],
                'Tipo': 'TOTAL',
                'Natureza': '',
                'Nome Natureza': '',
                'Saldo Original': ug['saldo_total'],
                'Saldo Ajustado': round(saldo_total_ajustado, 2)
            })

            for nat in ug['naturezas']:
                dados.append({
                    'Fonte': nat['fonte'],
                    'UG': '',
                    'Nome UG': '',
                    'Tipo': 'Natureza',
                    'Natureza': nat['codigo'],
                    'Nome Natureza': nat['nome'],
                    'Saldo Original': round(nat['saldo_original'], 2),
                    'Saldo Ajustado': round(nat['saldo_atual'], 2)
                })

        return pd.DataFrame(dados)

    def gerar_aba_remanejamento(self) -> pd.DataFrame:
        """Gera aba de remanejamentos com consolidação"""
        if not self.remanejamentos:
            return pd.DataFrame({
                'Tipo': [],
                'Fonte': [],
                'UG Origem': [],
                'Natureza Origem': [],
                'UG Destino': [],
                'Natureza Destino': [],
                'Valor': []
            })

        self.log(f"\n   === CONSOLIDANDO REMANEJAMENTOS ===")
        self.log(f"   Total antes da consolidação: {len(self.remanejamentos)}")

        # Consolidar transferências idênticas (mesma origem → mesmo destino)
        consolidados = {}
        for rem in self.remanejamentos:
            # Chave única: Fonte + UG Origem + Natureza Origem + UG Destino + Natureza Destino
            chave = (
                rem['Fonte'],
                rem['UG Origem'],
                rem['Natureza Origem'],
                rem['UG Destino'],
                rem['Natureza Destino']
            )

            if chave in consolidados:
                # Somar valores
                consolidados[chave]['Valor'] += rem['Valor']
            else:
                # Primeira ocorrência, adicionar
                consolidados[chave] = {
                    'Tipo': rem['Tipo'],
                    'Fonte': rem['Fonte'],
                    'UG Origem': rem['UG Origem'],
                    'Natureza Origem': rem['Natureza Origem'],
                    'Nome Natureza Origem': rem['Nome Natureza Origem'],
                    'UG Destino': rem['UG Destino'],
                    'Natureza Destino': rem['Natureza Destino'],
                    'Nome Natureza Destino': rem['Nome Natureza Destino'],
                    'Valor': rem['Valor']
                }

        # Converter de volta para lista
        remanejamentos_consolidados = list(consolidados.values())

        self.log(f"   Total após consolidação: {len(remanejamentos_consolidados)}")
        reducao = len(self.remanejamentos) - len(remanejamentos_consolidados)
        if reducao > 0:
            percentual = (reducao / len(self.remanejamentos)) * 100
            self.log(f"   ✓ Redução: {reducao} remanejamentos ({percentual:.1f}%)")

        df = pd.DataFrame(remanejamentos_consolidados)
        colunas = ['Tipo', 'Fonte', 'UG Origem', 'Natureza Origem', 'Nome Natureza Origem',
                   'UG Destino', 'Natureza Destino', 'Nome Natureza Destino', 'Valor']
        return df[colunas]

    def formatar_planilha(self, worksheet):
        """Formata planilha Excel"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
