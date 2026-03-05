import pandas as pd
import io
import os
from typing import Dict, List, Optional, Tuple
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class GeradorLote:
    """
    Gera arquivo Excel no formato de importação em lote do SIAFE a partir da aba
    'Remanejamentos' do arquivo de orçamento ajustado, usando as Regras de Mapeamento 41 e 100.
    """

    # Colunas do template SIAFE na ordem correta
    COLUNAS_SIAFE = [
        'Data Emissão', 'UG Emitente', 'Tipo Crédito', 'Origem Recursos',
        'UG Deduzida', 'UG Acrescida', 'Observação', 'Assunto', 'Processo',
        'Tipo Item', 'Órgão Orçamento', 'Unidade Orçamentária',
        'Programa de Trabalho', 'Fonte', 'Natureza', 'Autor Emenda',
        'Emenda Parlamentar', 'Território', 'Plano Orçamentário', 'Valor'
    ]

    # Valores fixos
    ASSUNTO = '4'               # Folha de Pessoal
    AUTOR_EMENDA = '0'
    EMENDA_PARLAMENTAR = '0000.E0000'
    TERRITORIO = 'TD0'

    def __init__(self):
        self.regra41 = None   # DataFrame da Regra 41
        self.regra100 = None  # DataFrame da Regra 100
        self.mapa_ug = {}     # Cache: UG -> {orgao, unidade, pt, plano}
        self.erros = []       # Erros encontrados durante o processamento

    # =========================================================================
    # FUNÇÕES DE FORMATAÇÃO
    # =========================================================================

    @staticmethod
    def formatar_unidade_orcamentaria(codigo: str) -> str:
        """
        Formata Unidade Orçamentária: XXXXX → XX.XXX
        Exemplo: 15000 → 15.000, 14102 → 14.102
        """
        codigo = str(codigo).strip().replace('.', '')
        if len(codigo) < 5:
            codigo = codigo.zfill(5)
        return f"{codigo[:2]}.{codigo[2:5]}"

    @staticmethod
    def formatar_fonte(fonte: int) -> str:
        """
        Formata Fonte: XXX → X.XX
        Exemplo: 500 → 5.00, 501 → 5.01, 761 → 7.61
        """
        fonte_str = str(int(fonte)).zfill(3)
        return f"{fonte_str[0]}.{fonte_str[1:3]}"

    @staticmethod
    def formatar_natureza(natureza: str) -> str:
        """
        Formata Natureza: XXXXXX → X.X.XX.XX
        Exemplo: 319011 → 3.1.90.11, 339036 → 3.3.90.36
        """
        nat = str(natureza).strip().replace('.', '')
        if len(nat) < 6:
            nat = nat.zfill(6)
        return f"{nat[0]}.{nat[1]}.{nat[2:4]}.{nat[4:6]}"

    @staticmethod
    def formatar_programa_trabalho(pt: str) -> str:
        """
        Formata Programa de Trabalho: XXXXXXXXXXXXXXXXXXX → X.XX.XXX.XX.XXX.XXXX.XXXX
        Tamanhos: [1, 2, 3, 2, 3, 4, 4] = 19 dígitos + 6 pontos = 25 chars
        Exemplo: 1491010618200091191 → 1.49.101.06.182.0009.1191

        Se já vier formatado (contém pontos), retorna como está.
        """
        if '.' in str(pt):
            return str(pt)
        pt_str = str(pt).strip()
        tamanhos = [1, 2, 3, 2, 3, 4, 4]
        partes = []
        pos = 0
        for tam in tamanhos:
            partes.append(pt_str[pos:pos + tam])
            pos += tam
        return '.'.join(partes)

    @staticmethod
    def formatar_valor(valor: float) -> str:
        """
        Formata valor no padrão brasileiro: 1234.56 → 1234,56
        """
        return f"{valor:.2f}".replace('.', ',')

    @staticmethod
    def formatar_ug(ug) -> str:
        """
        Garante que UG tenha 6 dígitos como string.
        """
        return str(int(ug)).zfill(6)

    # =========================================================================
    # CARREGAMENTO DAS REGRAS DE MAPEAMENTO
    # =========================================================================

    def carregar_regra41(self, caminho_ou_bytes):
        """
        Carrega a Regra de Mapeamento 41 (Programa de Trabalho).
        Estrutura:
          ORIGEM: Col 1=UG (tipo73), Col 2=tipo317, Col 3=tipo318, Col 4=tipo316
          DESTINO: Col 7=Órgão, Col 8=Unidade, Col 9=PT, Col 10=Plano
        """
        if isinstance(caminho_ou_bytes, str):
            self.regra41 = pd.read_excel(caminho_ou_bytes, header=None, engine='xlrd')
        else:
            self.regra41 = pd.read_excel(caminho_ou_bytes, header=None, engine='xlrd')

        self._construir_mapa_ug()

    def _construir_mapa_ug(self):
        """
        Constrói cache de mapeamento UG → dados do SIAFE a partir da Regra 41.
        Para UGs com múltiplos PTs, seleciona o PT terminando em '2500'
        (Administração da Unidade - padrão para Folha de Pessoal).
        """
        if self.regra41 is None:
            return

        # Agrupar por UG
        ug_rows = {}
        for i in range(2, len(self.regra41)):
            ug = str(self.regra41.iloc[i, 1]).strip()
            if ug == 'nan' or not ug:
                continue

            orgao = str(self.regra41.iloc[i, 7]).strip()
            unidade = str(self.regra41.iloc[i, 8]).strip()
            pt = str(self.regra41.iloc[i, 9]).strip()
            plano = str(self.regra41.iloc[i, 10]).strip()
            tipo317 = str(self.regra41.iloc[i, 2]).strip()
            tipo318 = str(self.regra41.iloc[i, 3]).strip()

            if orgao == 'nan' or pt == 'nan':
                continue

            if ug not in ug_rows:
                ug_rows[ug] = []
            ug_rows[ug].append({
                'orgao': orgao,
                'unidade': unidade,
                'pt': pt,
                'plano': plano if plano != 'nan' else '000001',
                'tipo317': tipo317,
                'tipo318': tipo318,
            })

        # Selecionar o melhor registro para cada UG
        for ug, rows in ug_rows.items():
            # Prioridade 1: PT terminando em '2500' com tipo317='001', tipo318='00'
            candidato = None
            for r in rows:
                if r['pt'].endswith('2500') and r['tipo317'] == '001' and r['tipo318'] == '00':
                    candidato = r
                    break

            # Prioridade 2: Qualquer PT terminando em '2500'
            if candidato is None:
                for r in rows:
                    if r['pt'].endswith('2500'):
                        candidato = r
                        break

            # Prioridade 3: Linha com tipo317='001', tipo318='00'
            if candidato is None:
                for r in rows:
                    if r['tipo317'] == '001' and r['tipo318'] == '00':
                        candidato = r
                        break

            # Prioridade 4: Primeira linha disponível
            if candidato is None:
                candidato = rows[0]

            self.mapa_ug[ug] = {
                'orgao': candidato['orgao'],
                'unidade': candidato['unidade'],
                'pt': candidato['pt'],
                'plano': candidato['plano'],
            }

    def carregar_regra100(self, caminho_ou_bytes):
        """
        Carrega a Regra de Mapeamento 100 (Fonte).
        Para nosso caso (fontes 500/501), a formatação é direta,
        mas o arquivo é carregado para validação futura.
        """
        if isinstance(caminho_ou_bytes, str):
            self.regra100 = pd.read_excel(caminho_ou_bytes, header=None, engine='xlrd')
        else:
            self.regra100 = pd.read_excel(caminho_ou_bytes, header=None, engine='xlrd')

    # =========================================================================
    # BUSCA NAS REGRAS
    # =========================================================================

    def buscar_dados_ug(self, ug: str) -> Optional[Dict]:
        """
        Busca Órgão, Unidade Orçamentária, Programa de Trabalho e Plano
        para uma UG usando o cache da Regra 41.
        """
        ug_str = str(int(float(ug))).zfill(6) if ug else ''
        if ug_str in self.mapa_ug:
            return self.mapa_ug[ug_str]
        self.erros.append(f"UG {ug_str} não encontrada na Regra 41")
        return None

    def obter_identificador_exercicio(self, fonte: int) -> str:
        """
        Retorna o Identificador de Exercício da Fonte.
        Baseado na análise da Regra 100: fontes 500 e 501 → '1' (Corrente).
        """
        # Todas as fontes 5.xx são Exercício Corrente (1)
        fonte_str = str(int(fonte)).zfill(3)
        if fonte_str[0] == '5':
            return '1'
        elif fonte_str[0] in ('6', '7', '8'):
            return '1'  # Baseado nos dados da Regra 100
        return '1'  # Default

    # =========================================================================
    # DETERMINAÇÃO DO TIPO DE CRÉDITO
    # =========================================================================

    def determinar_tipo_credito(self, ug_origem: str, ug_destino: str) -> Tuple[str, str]:
        """
        Determina o Tipo de Crédito e a Origem de Recursos baseado na foto real do SIAFE:
        - Mesma UG: Tipo '5' (Remanejamento Interno), Origem '0' (Não aplicável)
        - UGs diferentes: Tipo '1' (Suplementar), Origem '3' (Redução/Anulação de Dotação)

        Returns:
            Tuple (tipo_credito, origem_recursos)
        """
        ug_orig = str(int(float(ug_origem))).zfill(6)
        ug_dest = str(int(float(ug_destino))).zfill(6)
        if ug_orig == ug_dest:
            return '5', '0'   # Remanejamento Interno → Não aplicável
        else:
            return '1', '3'   # Suplementar → Redução/Anulação de Dotação

    # =========================================================================
    # GERAÇÃO DO ARQUIVO EM LOTE
    # =========================================================================

    def gerar_lote(
        self,
        df_remanejamentos: pd.DataFrame,
        data_emissao: str,
        observacao: str = '',
        processo: str = '',
        tipo_abertura: str = '1',
    ) -> Tuple[bytes, List[str]]:
        """
        Gera o arquivo Excel no formato de importação em lote do SIAFE.

        UG Emitente é derivada automaticamente da UG Acrescida (UG Destino) de cada linha.

        Cada remanejamento gera 2 linhas:
        - Linha 1 (Tipo Item = 2): REDUÇÃO na UG/Natureza Origem
        - Linha 2 (Tipo Item = 1): ACRÉSCIMO na UG/Natureza Destino

        Args:
            df_remanejamentos: DataFrame com a aba 'Remanejamentos'
            data_emissao: Data no formato DD/MM/AAAA
            ug_emitente: UG Emitente (6 dígitos)
            observacao: Texto de observação
            processo: Número do processo
            tipo_abertura: Tipo de Abertura (default '1' = Normal)

        Returns:
            Tuple com (bytes do Excel, lista de erros)
        """
        self.erros = []
        linhas = []

        for idx, row in df_remanejamentos.iterrows():
            ug_origem = row.get('UG Origem')
            ug_destino = row.get('UG Destino')
            nat_origem = row.get('Natureza Origem')
            nat_destino = row.get('Natureza Destino')
            valor = row.get('Valor', 0)
            fonte = row.get('Fonte')

            # Validações básicas
            if pd.isna(ug_origem) or pd.isna(ug_destino):
                self.erros.append(f"Linha {idx}: UG Origem ou Destino ausente")
                continue

            if pd.isna(valor) or valor <= 0:
                self.erros.append(f"Linha {idx}: Valor inválido ({valor})")
                continue

            # Formatar UGs
            ug_orig_fmt = self.formatar_ug(ug_origem)
            ug_dest_fmt = self.formatar_ug(ug_destino)

            # Determinar fonte
            fonte_str = str(fonte).strip() if not pd.isna(fonte) else ''
            if not fonte_str or fonte_str.lower() == 'nan':
                self.erros.append(
                    f"Linha {idx}: Fonte ausente — verifique os dados de origem. "
                    f"Linha ignorada (UG Origem={ug_origem}, UG Destino={ug_destino})."
                )
                continue
            try:
                fonte_int = int(float(fonte_str))
            except (ValueError, TypeError):
                self.erros.append(f"Linha {idx}: Fonte inválida '{fonte}' — linha ignorada.")
                continue

            # Tipo de Crédito e Origem de Recursos
            tipo_credito, origem_recursos = self.determinar_tipo_credito(ug_origem, ug_destino)

            # Dados da UG Origem (Regra 41)
            dados_ug_orig = self.buscar_dados_ug(ug_origem)
            if dados_ug_orig is None:
                dados_ug_orig = {
                    'orgao': '??', 'unidade': '??.???',
                    'pt': '?.??.???.??.???.????.????', 'plano': '000001'
                }

            # Dados da UG Destino (Regra 41)
            dados_ug_dest = self.buscar_dados_ug(ug_destino)
            if dados_ug_dest is None:
                dados_ug_dest = {
                    'orgao': '??', 'unidade': '??.???',
                    'pt': '?.??.???.??.???.????.????', 'plano': '000001'
                }

            # Formatar naturezas
            nat_orig_fmt = self.formatar_natureza(nat_origem)
            nat_dest_fmt = self.formatar_natureza(nat_destino)

            # Formatar fonte
            fonte_fmt = self.formatar_fonte(fonte_int)

            # Formatar valor
            valor_fmt = self.formatar_valor(valor)

            # UG Emitente = UG Acrescida (UG Destino) — confirmado na tela real do SIAFE
            # Campos compartilhados
            campos_base = {
                'Data Emissão': data_emissao,
                'UG Emitente': ug_dest_fmt,
                'Tipo Crédito': tipo_credito,
                'Origem Recursos': origem_recursos,
                'UG Deduzida': ug_orig_fmt,
                'UG Acrescida': ug_dest_fmt,
                'Observação': observacao,
                'Assunto': self.ASSUNTO,
                'Processo': processo,
                'Autor Emenda': self.AUTOR_EMENDA,
                'Emenda Parlamentar': self.EMENDA_PARLAMENTAR,
                'Território': self.TERRITORIO,
                'Valor': valor_fmt,
            }

            # LINHA 1: REDUÇÃO (Tipo Item = 2) - dados da UG Origem
            linha_reducao = dict(campos_base)
            linha_reducao['Tipo Item'] = '2'
            linha_reducao['Órgão Orçamento'] = dados_ug_orig['orgao']
            linha_reducao['Unidade Orçamentária'] = dados_ug_orig['unidade']
            linha_reducao['Programa de Trabalho'] = dados_ug_orig['pt']
            linha_reducao['Fonte'] = fonte_fmt
            linha_reducao['Natureza'] = nat_orig_fmt
            linha_reducao['Plano Orçamentário'] = dados_ug_orig['plano']
            linhas.append(linha_reducao)

            # LINHA 2: ACRÉSCIMO (Tipo Item = 1) - dados da UG Destino
            linha_acrescimo = dict(campos_base)
            linha_acrescimo['Tipo Item'] = '1'
            linha_acrescimo['Órgão Orçamento'] = dados_ug_dest['orgao']
            linha_acrescimo['Unidade Orçamentária'] = dados_ug_dest['unidade']
            linha_acrescimo['Programa de Trabalho'] = dados_ug_dest['pt']
            linha_acrescimo['Fonte'] = fonte_fmt
            linha_acrescimo['Natureza'] = nat_dest_fmt
            linha_acrescimo['Plano Orçamentário'] = dados_ug_dest['plano']
            linhas.append(linha_acrescimo)

        # Criar DataFrame final com todas as colunas como string
        df_siafe = pd.DataFrame(linhas, columns=self.COLUNAS_SIAFE)
        df_siafe = df_siafe.astype(str)

        # Gerar Excel usando openpyxl diretamente para preservar texto
        output = io.BytesIO()
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = 'SIAFE'

        # Escrever cabeçalho
        ws.append(list(df_siafe.columns))

        # Escrever dados como texto puro (evitar conversão numérica)
        for _, row in df_siafe.iterrows():
            ws.append([str(v) for v in row])

        # Formatar
        self._formatar_planilha(ws)

        wb.save(output)

        output.seek(0)
        return output.getvalue(), self.erros

    def _formatar_planilha(self, worksheet):
        """Formata a planilha com estilos visuais."""
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Larguras das colunas
        larguras = {
            'A': 14, 'B': 14, 'C': 12, 'D': 12, 'E': 14, 'F': 14,
            'G': 30, 'H': 10, 'I': 15, 'J': 10, 'K': 12, 'L': 14,
            'M': 30, 'N': 8, 'O': 12, 'P': 10, 'Q': 14, 'R': 8,
            'S': 14, 'T': 16
        }
        for col_letter, width in larguras.items():
            worksheet.column_dimensions[col_letter].width = width
